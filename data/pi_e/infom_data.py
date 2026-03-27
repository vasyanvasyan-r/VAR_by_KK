import pandas as pd
from pathlib import Path
from datetime import date, timedelta
import pickle
def find_gaps(sorted_list):
    gaps = {}
    end_of_df = sorted_list[-1]
    gap_id = None
    for i in range(1, len(sorted_list)):
        prev_num = sorted_list[i-1]
        curr_num = sorted_list[i]
        
        if curr_num - prev_num > 1:
            gap_start = prev_num + 1
            gap_end = curr_num - 1
            
            gap_id = f"gap_{len(gaps) + 1}"
            
            gaps[gap_id] = {
                'start': gap_start,
                'end': gap_end,
                'length': gap_end - gap_start + 1,
                'block': [gap_end+1, end_of_df] 
            }
        if gap_id is not None:
            gaps[gap_id]['block'][1] = curr_num
        
            
    return gaps

current_dir = Path.cwd() 
pi_e_folder = current_dir / 'data/pi_e'

files = [f for f in pi_e_folder.iterdir() if f.is_file()]

files_dict = {}
for file in files:
    fstr = file.name   
    files_dict['20' + fstr.split('.')[0].split('_')[-1]] = fstr

def get_month_range(start_year, start_month, end_year, end_month):
    current = date(start_year, start_month, 1)
    stop = date(end_year, end_month, 1)
    
    months_list = []
    while current <= stop:
        months_list.append(current.strftime('%Y-%m'))
        
        current = (current.replace(day=1) + timedelta(days=32)).replace(day=1)
        
    return months_list
start_y, start_m = 2020, 9   
end_y, end_m = 2026, 2      

months_list = get_month_range(start_y, start_m, end_y, end_m)

from openpyxl import load_workbook

for month in months_list:
    if month not in files_dict:
        continue

    filename = files_dict[month]
    file_path = pi_e_folder / filename
    
    try:
        # Открываем книгу только чтобы посмотреть имена листов
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Выбираем нужный лист
        if 'Таблица' in sheet_names:
            sheet_name = 'Таблица'
        elif 'Результат' in sheet_names:
            sheet_name = 'Результат'
        else:
            print(f'⚠️ В файле {filename} нет нужных листов. Доступны: {sheet_names}')
            files_dict[month] = None
            continue
            
        # Читаем данные уже зная имя листа
        files_dict[month] = pd.read_excel(file_path, sheet_name=sheet_name)
        
    except Exception as e:
        print(f'❌ Ошибка обработки файла {month}: {e}')
        files_dict[month] = None

pie_data = {}
for month in months_list:
    month_dict = {} 
    df = files_dict[month]
    if df is not None:
        df = df.set_axis(pd.RangeIndex(0, df.shape[1]), axis = 1)
        if df.iloc[:, 1].loc[['Население' in str(i) and 'в целом' in str(i) for i in df.iloc[:, 1]]].shape[0] != 0:
            start_index = df.iloc[:, 1].loc[['Население' in str(i) and 'в целом' in str(i) for i in df.iloc[:, 1]]].index.start
        elif df.iloc[:, 1].loc[['Все' in str(i) and 'опрошенные' in str(i) for i in df.iloc[:, 1]]].shape[0] != 0:
            start_index = df.iloc[:, 1].loc[['Все' in str(i) and 'опрошенные' in str(i) for i in df.iloc[:, 1]]].index.start
        else:
            print(f'бросил {month}')
            continue
        razrez_raw = df.loc[start_index].copy(deep = True)
        razrez_data = df.loc[start_index+1].copy(deep = True)
        razrez = razrez_raw.unique()[1:]
        razrez_indexes = razrez_raw.loc[[i in razrez for i in razrez_raw]].index.to_list().copy() + [df.shape[1]+1]
        config = {v :{'start': razrez_indexes[i],
                 'stop': razrez_indexes[i+1]-1, 
                 'data': razrez_data[razrez_indexes[i+1]:razrez_indexes[i+2]].to_list()} for i, v in enumerate(razrez[1:])}
        index_level_0 = ['Население в целом'] # Категория (Пол, Возраст...)
        index_level_1 = ['100%'] # Группа (мужчины, женщины...)

        for category, info in config.items():
            count = len(info['data'])
            # Добавляем название категории столько раз, сколько элементов в списке data
            index_level_0.extend([category.replace('-\n', '').replace('-', '')] * count)
            # Добавляем сами названия групп
            index_level_1.extend(info['data'])

        # 3. Создаем MultiIndex и присваиваем его DataFrame
        multi_index = pd.MultiIndex.from_arrays([index_level_0, index_level_1], names=['Категория', 'Группа'])

        index_list = []
        for i, row in df.iloc[start_index+2:].iterrows():
            if row.unique().shape[0] >= 3:
                index_list.append(i)
        gaps = find_gaps(index_list)
        print(f'========== {month} ==========')
        for gap in gaps:
            question = df.iloc[gaps[gap]['end'],0].replace('\xa0', ' ').replace('-\n', '')
            month_dict[question] = df.iloc[gaps[gap]['block'][0]:gaps[gap]['block'][1]+1].set_index(0).rename_axis(None).set_axis(multi_index, axis = 1).astype(float)
        pie_data[month] = month_dict
with open('data/pie_data.pkl', 'wb') as file:
    pickle.dump(pie_data, file)